from PageObjects import RegistrationPagePO
from Base import InitiateDriver
import pytest
import openpyxl

@pytest.fixture()
def setup():
    # Start browser on correct URL
    global driver
    global register

    driver = InitiateDriver.start_browser()
    register = RegistrationPagePO.RegistrationClass(driver)

    # Close browser
    yield
    InitiateDriver.close_browser()


def data_generator():
    # Would move to file for Data generator if doing this for real
    # wk = openpyxl.load_workbook("../Resources/test.doc")
    # sh = wk['Sheet1']
    # r = sh.max_row
    # li = []
    # li1 = []
    # for i in range(1, r+1):
    #     li1 = []
    #     un = sh.cell(i, 1)
    #     pw = sh.cell(i, 2)
    #     li1.insert(0, un.value)
    #     li1.insert(1, pw.value)
    #     li.insert(i-1, li1)

    li = [["uname1", "pass1"], ["uname2", "pass2"], ["uname3", "pass3"], ["uname4", "pass4"], ["uname5", "pass5"]]
    return li


@pytest.mark.parametrize('data', data_generator())
def test_verify_registration(setup, data):
    # Enter Data to the textbox
    # register = RegistrationPagePO.RegistrationClass(driver)
    register.enter_username(data[0])
    register.enter_password(data[1])